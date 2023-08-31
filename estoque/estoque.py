from openpyxl import load_workbook
import datetime
import pandas as pd

def cadastrarProduto():
    try:
        spreadsheet_name = 'estoque.xlsx'
        spreadsheet = load_workbook(spreadsheet_name)
        page_control = spreadsheet['estoque']
        id = int(input("Id do produto: "))
        name = input("nome do produto: ")
        estoque = int(input("estoque inicial: "))
        estoque_min = int(input("Estoque minimo: "))
        categoria = input("Categoria: ")
        page_control.append([id, name, estoque, estoque_min, categoria])
        spreadsheet.save(spreadsheet_name)
        print("produto cadastrado com sucesso")
    except Exception as e:
        print(e)


def add_entrada():
    try:
        spreadsheet_name = 'estoque.xlsx'
        spreadsheet = load_workbook(spreadsheet_name)
        logs= spreadsheet['logs']
        data = datetime.date.today()
        id_control = int(input("ID saída: "))
        type_E = 'E'
        cliente = input("Quem pegou o item: ")
        carro = input("Placa do Carro: ")
        modelo = input("Modelo do veículo: ")
        quantidade = input("Quantidade: ")
        id_item = int(input("ID do produto: "))
        
        new_row = [data, id_control, type_E, cliente, carro,
                   modelo, quantidade, id_item]

        # Adiciona uma nova linha à nova planilha
        logs.append(new_row)
        
        spreadsheet.save(spreadsheet_name)
        print('Entrada registrada')

    except Exception as e:
        print(e)

add_entrada()

def add_saida():
    try:
        spreadsheet_name = 'estoque.xlsx'
        spreadsheet = load_workbook(spreadsheet_name)
        logs= spreadsheet['logs']
        data = datetime.date.today()
        id_control = int(input("ID saída: "))
        type_E = 'S'
        cliente = input("Quem pegou o item: ")
        carro = input("Placa do Carro: ")
        modelo = input("Modelo do veículo: ")
        quantidade = input("Quantidade: ")
        id_item = int(input("ID do produto: "))
        
        new_row = [data, id_control, type_E, cliente, carro,
                   modelo, quantidade, id_item]

        # Adiciona uma nova linha à nova planilha
        logs.append(new_row)
        
        spreadsheet.save(spreadsheet_name)
        print('Saida registrada')

    except Exception as e:
        print(e)

add_saida()
