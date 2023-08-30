import openpyxl
import os


def create_spreadsheet():
    try:
        def data_columns_car():
            input_columns = input('Digite as colunas que você quer ter em Carros:\nobs separe por , sem espaços:\n').upper()
            print(input_columns)
            list_columns = []
            print(list_columns)
            columns = input_columns.split(',')
            print(columns)
            for data in columns:
                print(data)
                list_columns.append(data)
            columns_cars = list_columns
            print(columns_cars)
            return columns_cars

        def data_columns_os():
            input_columns = input('Digite as colunas que você quer ter em Ordem de serviço:\nobs separe por , sem espaços:\n').upper()
            list_columns = []
            columns = input_columns.split(',')
            for data in columns:
                list_columns.append(data)
            columns_os = list_columns
            return columns_os

        name_spreadsheet = input("nome da planilha: ")
        if os.path.isfile(f'{name_spreadsheet}.xlsx'):
            print("Uma planilha com esse nome já existe.")
            return
        spreadsheet = openpyxl.Workbook()
        spreadsheet.remove(spreadsheet.active)
        spreadsheet.create_sheet('Carros')
        spreadsheet.create_sheet('Ordem de Serviço')

        def page_car():
            name_sheet = 'Carros'
            spreadsheet_page = spreadsheet[name_sheet]
            spreadsheet_page.append(data_columns_car())

        def page_so():
            name_sheet = 'Ordem de Serviço'
            spreadsheet_page = spreadsheet[name_sheet]
            spreadsheet_page.append(data_columns_os())
        page_car()
        page_so()
        spreadsheet.save(f'{name_spreadsheet}.xlsx')
        print('Planilha criada com sucesso')
    except Exception as e:
        print(e)


def add_rows_car(data, models, brand, plate):
    try:
        name_spreadsheet = input("Nome da planilha: ")
        spreadsheet = openpyxl.load_workbook(f'{name_spreadsheet}.xlsx')
        page_car = spreadsheet['Carros']
        print('Cadastro do veiculo')
        page_car.append([data, models, brand, plate])
        spreadsheet.save(f'{name_spreadsheet}.xlsx')
        print('Carro Adicionado com sucesso')

    except NameError:
        print(NameError)


def add_rows_os(data, num_os, plate, value, revitalization,
                filter_ar, baterry, model, abner, fabiano,
                leandro, mazinho, vagner):
    try:
        name_spreadsheet = input('Nome da planilha')
        spreadsheet = openpyxl.load_workbook(f'{name_spreadsheet}.xlsx')
        page_os = spreadsheet['Ordem de Serviço']
        print('cadastro da ordem de serviço')
        page_os.append([data, num_os, plate, value, revitalization, filter_ar,
                        baterry, model, abner, fabiano, leandro, mazinho,
                        vagner])
        spreadsheet.save(f'{name_spreadsheet}.xlsx')
        print('OS Adicionado com Sucesso')
    except NameError:
        print('erro ao cadastrar os tente novamente')

# Dados do veiculo


def data_car():
    print('-----------')
    print('dados do veiculo')
    data_car = input("Data: ")
    model = input("Modelo do veiculo: ").upper()
    brad = input('Marca do veiculo: ').upper()
    plate_car = input('Placa do veiculo: ').upper()
    print('Dados Coletados com Sucesso')
    print("Adicionar Dados A planilha")
    add_rows_car(data_car, model, brad, plate_car)
    print('-----------')
# Dados da Os


def data_os():
    print('-----------')
    print('Dados da OS')
    data_os = input('Data: ')
    num_os = int(input('Numero da OS: '))
    plate_car = input('Placa do Veiculo: ')
    value_os = int(input('Valor da OS: '))
    revitalization = int(input('Valor da Revitalização do Motor: '))
    filter_cabine = int(input('Valor do filtro: '))
    baterry = int(input('Valor da Bateria: '))
    model_baterry = input('Modelo da Bateria: ').upper()
    abner = int(input("Valor Abner: "))
    fabiano = int(input("Valor Fabiano: "))
    leandro = int(input("Valor Leandro: "))
    mazinho = int(input("Valor Mazinho: "))
    vagner = int(input("Valor Vagner: "))
    print('Dados coletados com sucesso')
    print("Adicionar Dados A planilha")
    add_rows_os(data_os, num_os, plate_car, value_os, revitalization,
                filter_cabine, baterry, model_baterry, abner, fabiano,
                leandro, mazinho, vagner)
    print('-----------')
