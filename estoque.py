import datetime
from openpyxl import load_workbook


def form_create_product():
    id_product = int(input("Id do produto: "))
    name = input("nome do produto: ").upper()
    stock_initial = int(input("estoque inicial: "))
    stock_input = None
    output = None
    stock_actually = None
    stock_min = int(input("Estoque minimo: "))
    category = input("Categoria: ").upper()
    date = datetime.date.today()
    log_type = "Create"
    result = [id_product, name, stock_initial, stock_input, output,
              stock_actually, stock_min, category, date]
    result_log = [id_product, name, stock_initial, stock_input, output,
                  stock_actually, stock_min, category, date, log_type]
    print(result)
    return result, result_log


def form_output_product():
    data = datetime.date.today()
    id_control = int(input("ID do controle: "))
    type_E = 'S'
    cliente = input("Quem pegou o item: ").upper()
    carro = input("Modelo do Carro: ").upper()
    modelo = input("Placa do veículo: ").upper()
    quantidade = int(input("Quantidade: "))
    id_item = int(input("ID do produto: "))
    item = None
    result = [data, id_control, type_E, cliente, carro,
              modelo, quantidade, id_item]
    type_log = "Output"
    result_log = [data, id_control, type_E, cliente, carro,
                  modelo, quantidade, id_item, item, type_log]
    log = [data, id_control, type_E, cliente, carro, modelo, quantidade,
           id_item, item, type_log]
    return result, result_log, log


def form_input_product():
    data = datetime.date.today()
    id_control = int(input("ID do Controle: "))
    type_E = 'E'
    cliente = "ENTRADA"
    carro = "ENTRADA"
    modelo = "ENTRADA"
    quantidade = int(input("Quantidade: "))
    id_item = int(input("ID do produto: "))
    item = None
    type_log = "input"
    result = [data, id_control, type_E, cliente, carro,
              modelo, quantidade, id_item]
    result_log = [data, id_control, type_E, cliente, carro,
                  modelo, quantidade, id_item, item, type_log]
    log = [data, id_control, type_E, cliente, carro, modelo,
           quantidade, id_item, item, type_log]
    return result, result_log, log


def create_product():
    try:
        path = 'estoque.xlsx'
        workbook = load_workbook(path)
        page_control = workbook["logs_new_product"]
        logs_consult = workbook['logs']
        result, result_log = form_create_product()
        print(result, result_log)
        page_control.append(result)
        logs_consult.append(result_log)
        workbook.save(path)
        print("Produto Cadastrado com Sucesso")
    except Exception as e:
        print("Algo deu errado")
        print(e)


def input_products():
    try:
        path = 'estoque.xlsx'
        workbook = load_workbook(path)
        logs_control = workbook["logs_controle"]
        logs_input = workbook["logs_input"]
        logs = workbook["logs"]
        result, result_log, logs = form_input_product()
        # Adiciona uma nova linha à nova planilha
        logs_control.append(result)
        logs_input.append(result_log)
        logs.append(result_log)
        workbook.save(path)
        print('Produto Atualizado no Estoque')

    except Exception as e:
        print(e)


def output_product():
    try:
        xlsx_file_path = 'estoque.xlsx'
        workbook = load_workbook(xlsx_file_path)
        logs_control = workbook["logs_controle"]
        logs_output = workbook["logs_output"]
        logs = workbook["logs"]
        result, result_log, logs = form_output_product()
        logs_control.append(result)
        logs_output.append(result_log)
        logs.append(result_log)
        workbook.save(xlsx_file_path)
        print('registrando logs de saida')
        print("-------------------")
        print("Registrando Logs Geral")
        print("-------------------")
        print("Registrando Controle")
        print("-------------------")
        print(f'Saida registrada.....\nsaida do produto:\n{result}')

    except Exception as e:
        print(e)


def ver_pages():
    teste = load_workbook('estoque.xlsx')
    print(teste)
    for pag in teste:
        print(pag)


def menu_stock():
    print("--------------------------------------------------------------")
    action_estoque = int(input('escolha a opção\n1-Criar Produto\n2-Saida\n3-entrada\n4-Sair\n'))
    while action_estoque != 4:
        match action_estoque:
            case 1:
                create_product()
                print("Produto Cadastrado")
                action_estoque = int(input('escolha a opção\n1-Criar Produto\n2-Saida\n3-entrada\n4-Sair\n'))
                return action_estoque
            case 2:
                output_product()
                print("Estoque Atualizado")
                action_estoque = int(input('escolha a opção\n1-Criar Produto\n2-Saida\n3-entrada\n4-Sair\n'))
                return action_estoque
            case 3:
                input_products()
                print("Estoque Atualizado")
                action_estoque = int(input('escolha a opção\n1-Criar Produto\n2-Saida\n3-entrada\n4-Sair\n'))
                return action_estoque
            case 4:
                print("Voltando ao menu inicial")
                action = int(input('escolha a opção\n1-Criar Produto\n2-Saida\n3-entrada\n4-Sair\n'))
                return action_estoque
            case 5:
                ver_pages()
                action = int(input('escolha a opção\n1-Criar Produto\n2-Saida\n3-entrada\n4-Sair\n'))
        print("--------------------------------------------------------------")
        return action
